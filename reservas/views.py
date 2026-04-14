from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.utils import timezone
from decimal import Decimal
from django.db.models import Sum, Count, Q, Case, When, IntegerField, Value
from django.http import JsonResponse, HttpResponse
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
import uuid
import requests
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import Viaje, Reserva, HorarioViaje
from .forms import ViajeForm, PasajeroForm, HorarioViajeForm, AgregarPasajeroManualForm
from usuarios.decorators import coordinador_required, admin_required


def cupos_pagados_usuario(user):
    return Reserva.objects.filter(usuario=user, estado='pagado').aggregate(total=Sum('cantidad'))['total'] or 0


def calcular_monto_con_descuento(precio, cupos_base, cantidad):
    """Calcula el monto total aplicando descuentos cupo a cupo según posición histórica."""
    total = Decimal('0')
    for i in range(1, cantidad + 1):
        cupo_num = cupos_base + i
        if cupo_num % 10 == 0:       # gratis
            pass
        elif cupo_num % 5 == 0:      # 50% descuento
            total += precio * Decimal('0.5')
        else:
            total += precio
    return int(total)


def mensaje_descuento_rango(cupos_base, cantidad):
    """Retorna mensaje si algún cupo en el rango tiene descuento."""
    tiene_gratis = any((cupos_base + i) % 10 == 0 for i in range(1, cantidad + 1))
    tiene_50     = any((cupos_base + i) % 5  == 0 for i in range(1, cantidad + 1))
    if tiene_gratis:
        return '🎉 ¡Uno o más cupos de esta compra son GRATIS!'
    if tiene_50:
        return '🎊 ¡Uno o más cupos incluyen 50% de descuento!'
    return None


def enviar_comprobante_pago(reserva, monto_total):
    """Envía el comprobante de pago por correo al titular de cada reserva del grupo."""
    try:
        reservas_grupo = Reserva.objects.filter(
            grupo_compra=reserva.grupo_compra
        ).select_related('viaje__concierto', 'horario') if reserva.grupo_compra else [reserva]

        # Agrupar por email para no enviar duplicados
        enviados = set()
        for r in reservas_grupo:
            if not r.email or r.email in enviados:
                continue
            enviados.add(r.email)

            otras = [x for x in reservas_grupo if x.pk != r.pk]
            contexto = {
                'nombre_titular': r.nombre_titular or r.usuario.get_full_name() or r.usuario.username,
                'orden_compra':   r.orden_compra,
                'artista':        r.viaje.concierto.artista,
                'concierto':      r.viaje.concierto.nombre,
                'fecha_salida':   r.viaje.fecha_salida.strftime('%d/%m/%Y'),
                'cupos':          r.cantidad,
                'horario':        str(r.horario) if r.horario else '',
                'tipo_pasaje':    r.get_tipo_pasaje_display(),
                'monto_total':    f'{int(monto_total):,}'.replace(',', '.'),
                'reservas_adicionales': otras,
            }
            html = render_to_string('reservas/email_comprobante.html', contexto)
            msg = EmailMultiAlternatives(
                subject=f'Comprobante de pago — {r.viaje.concierto.artista}',
                body=f'Tu reserva para {r.viaje.concierto.artista} fue confirmada. Orden: {r.orden_compra}',
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[r.email],
            )
            msg.attach_alternative(html, 'text/html')
            msg.send(fail_silently=True)
    except Exception:
        pass  # No interrumpir el flujo si el correo falla


def es_coordinador(user):
    return user.is_authenticated and (
        user.is_superuser or
        user.groups.filter(name__in=['Administrador', 'Coordinador']).exists()
    )

def es_admin(user):
    return user.is_authenticated and (
        user.is_superuser or
        user.groups.filter(name='Administrador').exists()
    )


# ─── VIAJES PÚBLICOS ────────────────────────────────────────────────────────

def lista_viajes(request):
    viajes = Viaje.objects.filter(
        estado='disponible',
        fecha_salida__gte=timezone.now().date()
    ).select_related('concierto').order_by('fecha_salida')
    return render(request, 'reservas/lista_viajes.html', {'viajes': viajes})


@login_required
def detalle_viaje(request, pk):
    viaje = get_object_or_404(Viaje, pk=pk)
    perfil = request.user.perfilusuario if request.user.is_authenticated and hasattr(request.user, 'perfilusuario') else None
    reserva_pendiente = None
    cupos_pagados = 0
    if request.user.is_authenticated:
        reserva_pendiente = Reserva.objects.filter(
            viaje=viaje, usuario=request.user, estado='pendiente'
        ).first()
        cupos_pagados = Reserva.objects.filter(
            viaje=viaje, usuario=request.user, estado='pagado'
        ).aggregate(total=Sum('cantidad'))['total'] or 0
    cupos_base = 0
    proximo_hito = 0
    proximo_beneficio = ''
    if request.user.is_authenticated:
        cupos_base = cupos_pagados_usuario(request.user)
        ciclo = cupos_base % 10
        if ciclo < 4:
            proximo_hito = cupos_base + (4 - ciclo)
            proximo_beneficio = '50% descuento'
        elif ciclo < 9:
            proximo_hito = cupos_base + (9 - ciclo)
            proximo_beneficio = 'Reserva GRATIS'
        else:
            proximo_hito = cupos_base + (10 - ciclo)
            proximo_beneficio = '50% descuento'
    orden_ciudad = Case(
        When(ciudad='BELLOTO',      then=Value(1)),
        When(ciudad='VIÑA DEL MAR', then=Value(2)),
        When(ciudad='VALPARAÍSO',   then=Value(3)),
        When(ciudad='PLACILLA',     then=Value(4)),
        When(ciudad='CASABLANCA',   then=Value(5)),
        default=Value(99),
        output_field=IntegerField(),
    )
    horarios = viaje.horarios.annotate(orden_ciudad=orden_ciudad).order_by('orden_ciudad', 'salida', 'hora_salida')
    context = {
        'viaje': viaje,
        'horarios': horarios,
        'reserva_pendiente': reserva_pendiente,
        'cupos_pagados': cupos_pagados,
        'perfil': perfil,
        'cupos_base': cupos_base,
        'proximo_hito': proximo_hito,
        'proximo_beneficio': proximo_beneficio,
    }
    return render(request, 'reservas/detalle_viaje.html', context)


# ─── WEBPAY / TRANSBANK ──────────────────────────────────────────────────────

def _precio_para_tipo(horario, tipo_pasaje, viaje=None):
    """Devuelve el precio unitario según el tipo de pasaje y el horario seleccionado."""
    if tipo_pasaje == 'solo_vuelta':
        # El precio de solo vuelta viene del viaje, no del horario
        if viaje is not None:
            return viaje.precio_vuelta
        return 0
    if horario is None:
        return 0
    if tipo_pasaje == 'solo_ida':
        return horario.precio_ida
    return horario.precio  # ida_vuelta


@login_required
def reservar_pendiente(request, pk):
    if request.method != 'POST':
        return redirect('detalle_viaje', pk=pk)

    viaje = get_object_or_404(Viaje, pk=pk, estado='disponible')

    # Leer ítems del carrito (JSON)
    try:
        items_raw = request.POST.get('carrito_items', '')
        items = json.loads(items_raw) if items_raw else []
    except (ValueError, TypeError):
        items = []

    # Fallback: un solo ítem
    if not items:
        try:
            cantidad = max(1, int(request.POST.get('cantidad', 1)))
        except (ValueError, TypeError):
            cantidad = 1
        items = [{
            'tipo_pasaje': request.POST.get('tipo_pasaje', 'ida_vuelta'),
            'horario_id': request.POST.get('horario_id', ''),
            'cantidad': cantidad,
        }]

    if viaje.cupos_disponibles <= 0:
        messages.error(request, 'Lo sentimos, este viaje ya no tiene cupos disponibles.')
        return redirect('detalle_viaje', pk=pk)

    cupos_base_pend = cupos_pagados_usuario(request.user)
    grupo = f'BC-{uuid.uuid4().hex[:12].upper()}'

    # Eliminar reservas pendientes previas
    Reserva.objects.filter(viaje=viaje, usuario=request.user, estado='pendiente').delete()

    cupos_acumulados = cupos_base_pend
    for item in items:
        try:
            cantidad_item = max(1, int(item.get('cantidad', 1)))
        except (ValueError, TypeError):
            cantidad_item = 1
        tipo_pasaje = item.get('tipo_pasaje', 'ida_vuelta')
        horario_id = item.get('horario_id', '')
        rut_titular = item.get('rut', '').strip()
        nombre_titular = item.get('nombre_titular', '').strip()
        ciudad_vuelta = item.get('ciudad_vuelta', '').strip()
        contacto = item.get('contacto', '').strip()
        email_titular = item.get('email', '').strip()
        horario = None
        if horario_id:
            try:
                horario = HorarioViaje.objects.get(pk=horario_id, viaje=viaje)
            except HorarioViaje.DoesNotExist:
                pass

        precio_unitario = _precio_para_tipo(horario, tipo_pasaje, viaje)
        monto_item = calcular_monto_con_descuento(precio_unitario, cupos_acumulados, cantidad_item)
        cupos_acumulados += cantidad_item

        Reserva.objects.create(
            viaje=viaje,
            horario=horario,
            usuario=request.user,
            estado='pendiente',
            cantidad=cantidad_item,
            tipo_pasaje=tipo_pasaje,
            monto=monto_item,
            grupo_compra=grupo,
            rut=rut_titular,
            nombre_titular=nombre_titular,
            ciudad_vuelta=ciudad_vuelta,
            contacto=contacto,
            email=email_titular,
        )

    messages.success(request, 'Reserva realizada. Recuerda pagar antes de 5 días previos al evento.')
    return redirect('mis_reservas')


@login_required
def iniciar_pago(request, pk):
    viaje = get_object_or_404(Viaje, pk=pk, estado='disponible')

    # Leer ítems del carrito (JSON)
    try:
        items_raw = request.POST.get('carrito_items', '')
        items = json.loads(items_raw) if items_raw else []
    except (ValueError, TypeError):
        items = []

    # Fallback: un solo ítem (compatibilidad)
    if not items:
        try:
            cantidad = max(1, int(request.POST.get('cantidad', 1)))
        except (ValueError, TypeError):
            cantidad = 1
        items = [{
            'tipo_pasaje': request.POST.get('tipo_pasaje', 'ida_vuelta'),
            'horario_id': request.POST.get('horario_id', ''),
            'cantidad': cantidad,
        }]

    if viaje.cupos_disponibles <= 0:
        messages.error(request, 'Lo sentimos, este viaje ya no tiene cupos disponibles.')
        return redirect('detalle_viaje', pk=pk)

    cupos_base_pago = cupos_pagados_usuario(request.user)
    grupo = f'BC-{uuid.uuid4().hex[:12].upper()}'

    # Eliminar reservas pendientes previas de este usuario en este viaje
    Reserva.objects.filter(viaje=viaje, usuario=request.user, estado='pendiente').delete()

    reservas_creadas = []
    monto_total = 0
    cupos_acumulados = cupos_base_pago

    for item in items:
        try:
            cantidad_item = max(1, int(item.get('cantidad', 1)))
        except (ValueError, TypeError):
            cantidad_item = 1
        tipo_pasaje = item.get('tipo_pasaje', 'ida_vuelta')
        horario_id = item.get('horario_id', '')
        rut_titular = item.get('rut', '').strip()
        nombre_titular = item.get('nombre_titular', '').strip()
        ciudad_vuelta = item.get('ciudad_vuelta', '').strip()
        contacto = item.get('contacto', '').strip()
        email_titular = item.get('email', '').strip()
        horario = None
        if horario_id:
            try:
                horario = HorarioViaje.objects.get(pk=horario_id, viaje=viaje)
            except HorarioViaje.DoesNotExist:
                pass

        precio_unitario = _precio_para_tipo(horario, tipo_pasaje, viaje)
        monto_item = calcular_monto_con_descuento(precio_unitario, cupos_acumulados, cantidad_item)
        cupos_acumulados += cantidad_item
        monto_total += monto_item

        r = Reserva.objects.create(
            viaje=viaje,
            horario=horario,
            usuario=request.user,
            estado='pendiente',
            cantidad=cantidad_item,
            tipo_pasaje=tipo_pasaje,
            monto=monto_item,
            grupo_compra=grupo,
            rut=rut_titular,
            nombre_titular=nombre_titular,
            ciudad_vuelta=ciudad_vuelta,
            contacto=contacto,
            email=email_titular,
        )
        reservas_creadas.append(r)

    if not reservas_creadas:
        messages.error(request, 'El carrito está vacío.')
        return redirect('detalle_viaje', pk=pk)

    # Reserva principal (primera) para WebPay
    reserva_principal = reservas_creadas[0]

    # Si el monto total es 0 (cupos gratis), confirmar sin WebPay
    if monto_total == 0:
        Reserva.objects.filter(grupo_compra=grupo).update(estado='pagado')
        messages.success(request, '¡Reserva confirmada! Tu cupo fue GRATIS por tu fidelidad. 🎉')
        return redirect('mis_reservas')

    # Iniciar transacción Webpay con monto total del carrito
    url_retorno = request.build_absolute_uri(f'/reservas/webpay/retorno/{reserva_principal.id}/')
    headers = {
        'Tbk-Api-Key-Id': settings.TRANSBANK_COMMERCE_CODE,
        'Tbk-Api-Key-Secret': settings.TRANSBANK_API_KEY,
        'Content-Type': 'application/json',
    }
    base_url = 'https://webpay3gint.transbank.cl' if settings.TRANSBANK_ENVIRONMENT == 'TEST' else 'https://webpay3g.transbank.cl'

    payload = {
        'buy_order': grupo,
        'session_id': f'session-{request.user.id}-{grupo}',
        'amount': int(monto_total),
        'return_url': url_retorno,
    }

    try:
        resp = requests.post(
            f'{base_url}/rswebpaytransaction/api/webpay/v1.2/transactions',
            headers=headers,
            json=payload,
            timeout=15
        )
        data = resp.json()
        if 'token' in data:
            reserva_principal.token_webpay = data['token']
            reserva_principal.save()
            return render(request, 'reservas/webpay_redirect.html', {
                'url': data['url'],
                'token': data['token'],
            })
        else:
            Reserva.objects.filter(grupo_compra=grupo).delete()
            messages.error(request, 'Error al conectar con Webpay. Intenta nuevamente.')
    except Exception as e:
        Reserva.objects.filter(grupo_compra=grupo).delete()
        messages.error(request, f'Error de conexión con Transbank: {str(e)}')

    return redirect('detalle_viaje', pk=pk)


@login_required
def retorno_webpay(request, reserva_id):
    reserva = get_object_or_404(Reserva, id=reserva_id, usuario=request.user)
    token = request.POST.get('token_ws') or request.GET.get('token_ws')

    if not token:
        reserva.estado = 'cancelado'
        reserva.save()
        messages.error(request, 'Pago cancelado o no completado.')
        return redirect('lista_viajes')

    headers = {
        'Tbk-Api-Key-Id': settings.TRANSBANK_COMMERCE_CODE,
        'Tbk-Api-Key-Secret': settings.TRANSBANK_API_KEY,
        'Content-Type': 'application/json',
    }
    if settings.TRANSBANK_ENVIRONMENT == 'TEST':
        base_url = 'https://webpay3gint.transbank.cl'
    else:
        base_url = 'https://webpay3g.transbank.cl'

    try:
        resp = requests.put(
            f'{base_url}/rswebpaytransaction/api/webpay/v1.2/transactions/{token}',
            headers=headers,
            timeout=15
        )
        data = resp.json()
        if data.get('response_code') == 0:
            # Marcar todos los ítems del grupo como pagados
            if reserva.grupo_compra:
                Reserva.objects.filter(grupo_compra=reserva.grupo_compra).update(estado='pagado', pago='WEB')
            else:
                reserva.estado = 'pagado'
                reserva.pago = 'WEB'
                reserva.save()
            # Actualizar estado del viaje si está lleno
            if reserva.viaje.cupos_disponibles <= 0:
                reserva.viaje.estado = 'completo'
                reserva.viaje.save()
            monto_total_pagado = Reserva.objects.filter(
                grupo_compra=reserva.grupo_compra
            ).aggregate(total=Sum('monto'))['total'] if reserva.grupo_compra else reserva.monto
            enviar_comprobante_pago(reserva, monto_total_pagado)
            messages.success(request, '¡Reserva confirmada! Tu pago fue procesado exitosamente.')
            return render(request, 'reservas/pago_exitoso.html', {
                'reserva': reserva,
                'data': data,
                'monto_total': monto_total_pagado,
            })
        else:
            if reserva.grupo_compra:
                Reserva.objects.filter(grupo_compra=reserva.grupo_compra).update(estado='cancelado')
            else:
                reserva.estado = 'cancelado'
                reserva.save()
            messages.error(request, 'El pago fue rechazado por Webpay.')
    except Exception as e:
        messages.error(request, f'Error al confirmar el pago: {str(e)}')

    return render(request, 'reservas/pago_fallido.html', {'reserva': reserva})


# ─── MIS RESERVAS ────────────────────────────────────────────────────────────

@login_required
def mis_reservas(request):
    reservas = Reserva.objects.filter(usuario=request.user).select_related('viaje__concierto')
    return render(request, 'reservas/mis_reservas.html', {'reservas': reservas})


# ─── GESTIÓN COORDINADOR ─────────────────────────────────────────────────────

@login_required
@user_passes_test(es_coordinador)
def gestion_viajes(request):
    hoy = timezone.now().date()
    # Marcar como realizados todos los viajes pasados que aún no lo estén
    Viaje.objects.filter(fecha_salida__lt=hoy).exclude(estado='realizado').update(estado='realizado')
    if es_admin(request.user):
        viajes = Viaje.objects.filter(fecha_salida__gte=hoy).select_related('concierto', 'coordinador')
        viajes_pasados = Viaje.objects.filter(fecha_salida__lt=hoy).select_related('concierto', 'coordinador').order_by('-fecha_salida')
    else:
        viajes = Viaje.objects.filter(coordinador=request.user, fecha_salida__gte=hoy).select_related('concierto')
        viajes_pasados = Viaje.objects.filter(coordinador=request.user, fecha_salida__lt=hoy).select_related('concierto').order_by('-fecha_salida')
    return render(request, 'reservas/gestion_viajes.html', {'viajes': viajes, 'viajes_pasados': viajes_pasados})


@login_required
@user_passes_test(es_coordinador)
def crear_viaje(request):
    from core.models import Concierto
    from django.http import JsonResponse
    from django.urls import reverse
    if request.method == 'POST':
        form = ViajeForm(request.POST)
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if form.is_valid():
            texto = form.cleaned_data['concierto_nombre'].strip()
            concierto = Concierto.objects.filter(
                Q(artista__icontains=texto) | Q(nombre__icontains=texto)
            ).first()
            if not concierto:
                form.add_error('concierto_nombre', f'No se encontró ningún concierto con "{texto}". Créalo primero en el Calendario.')
                if is_ajax:
                    return JsonResponse({'ok': False, 'errors': {'concierto_nombre': [f'No se encontró ningún concierto con "{texto}". Créalo primero en el Calendario.']}})
            else:
                viaje = form.save(commit=False)
                viaje.concierto = concierto
                viaje.coordinador = request.user
                viaje.save()
                if is_ajax:
                    return JsonResponse({'ok': True, 'redirect': reverse('horarios_viaje', args=[viaje.pk])})
                messages.success(request, 'Viaje creado. Ahora agrega los horarios.')
                return redirect('horarios_viaje', pk=viaje.pk)
        else:
            if is_ajax:
                errors = {field: list(errs) for field, errs in form.errors.items()}
                return JsonResponse({'ok': False, 'errors': errors})
    else:
        form = ViajeForm()
    return render(request, 'reservas/form_viaje.html', {'form': form, 'titulo': 'Crear Nuevo Viaje'})


@login_required
@user_passes_test(es_coordinador)
def editar_viaje(request, pk):
    from core.models import Concierto
    from django.http import JsonResponse
    viaje = get_object_or_404(Viaje, pk=pk)
    if request.method == 'POST':
        form = ViajeForm(request.POST, instance=viaje)
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if form.is_valid():
            texto = form.cleaned_data['concierto_nombre'].strip()
            concierto = Concierto.objects.filter(
                Q(artista__icontains=texto) | Q(nombre__icontains=texto)
            ).first()
            if not concierto:
                form.add_error('concierto_nombre', f'No se encontró ningún concierto con "{texto}". Créalo primero en el Calendario.')
                if is_ajax:
                    return JsonResponse({'ok': False, 'errors': {'concierto_nombre': [f'No se encontró ningún concierto con "{texto}". Créalo primero en el Calendario.']}})
            else:
                viaje = form.save(commit=False)
                viaje.concierto = concierto
                viaje.save()
                if is_ajax:
                    return JsonResponse({'ok': True})
                messages.success(request, 'Viaje actualizado exitosamente.')
                return redirect('gestion_viajes')
        else:
            if is_ajax:
                errors = {field: list(errs) for field, errs in form.errors.items()}
                return JsonResponse({'ok': False, 'errors': errors})
    else:
        form = ViajeForm(instance=viaje)
    return render(request, 'reservas/form_viaje.html', {'form': form, 'titulo': 'Editar Viaje', 'viaje': viaje})


@login_required
@user_passes_test(es_coordinador)
def eliminar_viaje(request, pk):
    viaje = get_object_or_404(Viaje, pk=pk)
    if request.method == 'POST':
        viaje.delete()
        messages.success(request, 'Viaje eliminado.')
        return redirect('gestion_viajes')
    return render(request, 'reservas/confirmar_eliminar.html', {'objeto': viaje, 'tipo': 'viaje'})


@login_required
@user_passes_test(es_coordinador)
def pasajeros_viaje(request, pk):
    viaje = get_object_or_404(Viaje, pk=pk)
    pasajeros = Reserva.objects.filter(viaje=viaje).exclude(estado='cancelado').select_related('usuario__perfilusuario', 'horario').order_by('fecha_reserva')
    form = AgregarPasajeroManualForm(viaje=viaje)
    return render(request, 'reservas/pasajeros_viaje.html', {'viaje': viaje, 'pasajeros': pasajeros, 'form': form})


@login_required
@user_passes_test(es_coordinador)
def agregar_pasajero_manual(request, pk):
    viaje = get_object_or_404(Viaje, pk=pk)
    if request.method == 'POST':
        form = AgregarPasajeroManualForm(request.POST, viaje=viaje)
        if form.is_valid():
            reserva = form.save(commit=False)
            reserva.viaje = viaje
            reserva.usuario = request.user
            reserva.estado = 'pagado'
            reserva.pago = 'TRANSFERENCIA'
            if reserva.contacto and not reserva.contacto.startswith('+56'):
                reserva.contacto = f'+56 9 {reserva.contacto}'
            reserva.save()
            if reserva.email:
                enviar_comprobante_pago(reserva, reserva.monto)
            messages.success(request, 'Pasajero agregado correctamente.')
        else:
            messages.error(request, 'Error al agregar el pasajero. Revisa los datos.')
    return redirect('pasajeros_viaje', pk=pk)


@login_required
@user_passes_test(es_coordinador)
def exportar_pasajeros_excel(request, pk):
    viaje = get_object_or_404(Viaje, pk=pk)
    reservas_qs = Reserva.objects.filter(viaje=viaje).exclude(estado='cancelado').select_related('usuario__perfilusuario', 'horario').order_by('fecha_reserva')

    import re
    import io
    import xlsxwriter

    TIPO_LABELS = {
        'ida_vuelta':  'Ida y Vuelta',
        'solo_ida':    'Solo Ida',
        'solo_vuelta': 'Solo Vuelta',
    }

    output = io.BytesIO()
    wb = xlsxwriter.Workbook(output, {'in_memory': True})
    ws = wb.add_worksheet('Pasajeros')

    fmt_header = wb.add_format({
        'bold': True, 'bg_color': '#FFD600', 'font_color': '#000000',
        'align': 'center', 'valign': 'vcenter', 'border': 1,
    })
    fmt_center = wb.add_format({'align': 'center', 'valign': 'vcenter'})
    fmt_text   = wb.add_format({'valign': 'vcenter'})

    encabezados = ['Titular', 'Cupos', 'Contacto', 'Origen', 'Retorno', 'Tipo de Pasaje']
    anchos      = [30,        8,       18,          25,       25,        18]

    for col, (texto, ancho) in enumerate(zip(encabezados, anchos)):
        ws.write(0, col, texto, fmt_header)
        ws.set_column(col, col, ancho)

    ws.set_row(0, 18)

    for i, r in enumerate(reservas_qs):
        row = i + 1
        origen = f"{r.horario.ciudad} {r.horario.hora_salida.strftime('%H:%M')} hrs" if r.horario else '—'
        ws.write(row, 0, r.nombre_titular or '—',              fmt_text)
        ws.write(row, 1, r.cantidad,                           fmt_center)
        ws.write(row, 2, r.contacto or '—',                    fmt_text)
        ws.write(row, 3, origen,                               fmt_text)
        ws.write(row, 4, (r.ciudad_vuelta or '—').upper(),     fmt_text)
        ws.write(row, 5, TIPO_LABELS.get(r.tipo_pasaje, '—'),  fmt_text)

    wb.close()

    artista_seguro = re.sub(r'[^\w\-]', '_', viaje.concierto.artista)
    nombre_archivo = f"pasajeros_{artista_seguro}_{viaje.fecha_salida.strftime('%Y%m%d')}.xlsx"

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response


@login_required
@user_passes_test(es_coordinador)
def editar_pasajero(request, pk):
    reserva = get_object_or_404(Reserva, pk=pk)
    if request.method == 'POST':
        form = PasajeroForm(request.POST, instance=reserva)
        if form.is_valid():
            form.save()
            messages.success(request, 'Reserva del pasajero actualizada.')
            return redirect('pasajeros_viaje', pk=reserva.viaje.pk)
    else:
        form = PasajeroForm(instance=reserva)
    return render(request, 'reservas/form_pasajero.html', {'form': form, 'reserva': reserva})


@login_required
@user_passes_test(es_coordinador)
def editar_pasajero_manual(request, pk):
    reserva = get_object_or_404(Reserva, pk=pk)
    if request.method == 'POST':
        form = AgregarPasajeroManualForm(request.POST, instance=reserva, viaje=reserva.viaje)
        if form.is_valid():
            reserva = form.save(commit=False)
            if reserva.contacto and not reserva.contacto.startswith('+56'):
                reserva.contacto = f'+56 9 {reserva.contacto}'
            reserva.save()
            messages.success(request, 'Pasajero actualizado correctamente.')
        else:
            messages.error(request, 'Error al actualizar el pasajero. Revisa los datos.')
    return redirect('pasajeros_viaje', pk=reserva.viaje.pk)


@login_required
@user_passes_test(es_coordinador)
def eliminar_pasajero(request, pk):
    reserva = get_object_or_404(Reserva, pk=pk)
    viaje_pk = reserva.viaje.pk
    if request.method == 'POST':
        reserva.estado = 'cancelado'
        reserva.save()
        messages.success(request, 'Pasajero eliminado del viaje.')
        return redirect('pasajeros_viaje', pk=viaje_pk)
    return render(request, 'reservas/confirmar_eliminar.html', {'objeto': reserva, 'tipo': 'pasajero'})


# ─── HORARIOS DE VIAJE ───────────────────────────────────────────────────────

@login_required
@user_passes_test(es_coordinador)
def horarios_viaje(request, pk):
    viaje = get_object_or_404(Viaje, pk=pk)
    horarios = viaje.horarios.all()
    form = HorarioViajeForm()
    if request.method == 'POST':
        form = HorarioViajeForm(request.POST)
        if form.is_valid():
            horario = form.save(commit=False)
            horario.viaje = viaje
            horario.save()
            messages.success(request, 'Horario agregado.')
            return redirect('horarios_viaje', pk=pk)
    return render(request, 'reservas/horarios_viaje.html', {'viaje': viaje, 'horarios': horarios, 'form': form})


@login_required
@user_passes_test(es_coordinador)
def editar_horario(request, pk):
    horario = get_object_or_404(HorarioViaje, pk=pk)
    if request.method == 'POST':
        form = HorarioViajeForm(request.POST, instance=horario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Horario actualizado.')
            return redirect('horarios_viaje', pk=horario.viaje.pk)
    else:
        form = HorarioViajeForm(instance=horario)
    return render(request, 'reservas/horarios_viaje.html', {
        'viaje': horario.viaje,
        'horarios': horario.viaje.horarios.all(),
        'form': form,
        'editando': horario,
    })


@login_required
@user_passes_test(es_coordinador)
def eliminar_horario(request, pk):
    horario = get_object_or_404(HorarioViaje, pk=pk)
    viaje_pk = horario.viaje.pk
    if request.method == 'POST':
        horario.delete()
        messages.success(request, 'Horario eliminado.')
        return redirect('horarios_viaje', pk=viaje_pk)
    return render(request, 'reservas/confirmar_eliminar.html', {'objeto': horario, 'tipo': 'horario'})


# ─── MÓDULO AUDITORÍA (ADMIN) ─────────────────────────────────────────────────

@login_required
@user_passes_test(es_admin)
def auditoria_ganancias(request):
    viajes = Viaje.objects.filter(estado='disponible').select_related('concierto', 'coordinador').prefetch_related('reserva_set').order_by('-fecha_salida')
    viajes_finalizados = Viaje.objects.exclude(estado='disponible').select_related('concierto', 'coordinador').prefetch_related('reserva_set').order_by('-fecha_salida')
    total_general = sum(v.ganancia_total for v in viajes)
    total_pasajeros = sum(v.cupos_ocupados for v in viajes) + sum(v.cupos_ocupados for v in viajes_finalizados)
    total_finalizados = sum(v.ganancia_total for v in viajes_finalizados)
    total_combinado = total_general + total_finalizados
    context = {
        'viajes': viajes,
        'viajes_finalizados': viajes_finalizados,
        'total_general': total_general,
        'total_pasajeros': total_pasajeros,
        'total_finalizados': total_finalizados,
        'total_combinado': total_combinado,
    }
    return render(request, 'reservas/auditoria.html', context)


@login_required
@user_passes_test(es_admin)
def exportar_auditoria_excel(request):
    mes  = request.GET.get('mes', '')   # formato YYYY-MM
    anio = request.GET.get('anio', '')  # formato YYYY

    viajes_curso = Viaje.objects.filter(estado='disponible').select_related('concierto', 'coordinador').order_by('-fecha_salida')
    viajes_fin   = Viaje.objects.exclude(estado='disponible').select_related('concierto', 'coordinador').order_by('-fecha_salida')

    def filtrar(qs):
        if mes and mes != 'todos':
            try:
                anio_m, num_m = mes.split('-')
                qs = qs.filter(fecha_salida__year=anio_m, fecha_salida__month=num_m)
            except ValueError:
                pass
        elif anio and anio != 'todos':
            qs = qs.filter(fecha_salida__year=anio)
        return qs

    viajes_curso = filtrar(viajes_curso)
    viajes_fin   = filtrar(viajes_fin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Auditoría'

    amarillo    = PatternFill('solid', fgColor='FFD600')
    font_bold   = Font(bold=True, color='111111')
    font_total  = Font(bold=True, color='FFFFFF')
    fill_total  = PatternFill('solid', fgColor='111111')
    center      = Alignment(horizontal='center', vertical='center')

    cabeceras = ['Artista', 'Concierto', 'Fecha Salida', 'Origen', 'Destino', 'Cupos Totales', 'Pasajeros Pagados', 'Ganancia (CLP)']
    ws.append(cabeceras)
    for cell in ws[1]:
        cell.fill      = amarillo
        cell.font      = font_bold
        cell.alignment = center

    def escribir_filas(qs, etiqueta):
        for v in qs:
            ws.append([
                v.concierto.artista,
                v.concierto.nombre,
                v.fecha_salida.strftime('%d/%m/%Y'),
                v.origen,
                v.destino,
                v.cupos_totales,
                v.cupos_ocupados,
                int(v.ganancia_total),
            ])

    escribir_filas(viajes_curso, 'En Curso')
    escribir_filas(viajes_fin,   'Finalizado')

    # Fila de total
    total = sum(int(v.ganancia_total) for v in list(viajes_curso) + list(viajes_fin))
    fila_total = ws.max_row + 1
    ws.cell(fila_total, 7, 'TOTAL')
    ws.cell(fila_total, 8, total)
    for col in range(1, 9):
        c = ws.cell(fila_total, col)
        c.fill = fill_total
        c.font = font_total
        c.alignment = center

    anchos = [22, 22, 14, 20, 20, 12, 14, 18]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    nombre_archivo = f'auditoria_{mes or anio or "completa"}.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


# ─── COMPRAS ────────────────────────────────────────────────────────────────
@login_required
@user_passes_test(es_coordinador)
def compras(request):
    reservas = (
        Reserva.objects
        .select_related('viaje__concierto', 'usuario')
        .order_by('-fecha_reserva')
    )

    from django.db.models import Sum
    context = {
        'reservas': reservas,
        'total': reservas.count(),
        'total_cupos': reservas.exclude(estado='cancelado').aggregate(total=Sum('cantidad'))['total'] or 0,
    }
    return render(request, 'reservas/compras.html', context)
